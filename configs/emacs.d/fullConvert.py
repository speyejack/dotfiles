import os
import re
from openpyxl import load_workbook
from autoOrgConverter import *

def createPackage(filename):
    try:
        fileStr = getFileStr(filename)
        package = getPackage(fileStr)
    except Exception as e:
        print("Package: " + filename)
        raise e
    return package

blacklist = ["init-elpa.el", "user-config.el", "init-use-package.el"]

def collectAllPackages(base):
    packages_collect = {}
    for root, dirnames, filenames in os.walk(base):
        filenames = filter(lambda x: ".el" in x, filenames)
        filenames = filter(lambda x: "layer" not in x, filenames)
        filenames = filter(lambda x: x not in blacklist, filenames)
        packages = {pack["name"]: pack for pack in [createPackage(root + "/" + filename) for filename in filenames]}
        packages_collect.update(packages)
    return packages_collect

def getChangedPackages(packages):
    changed = set()
    for name in packages:
        package = packages[name]
        keywords = package['keywords']
        if "init" in keywords or "config" in keywords:
            changed.add(name)
        
    return changed

def getTopHeaders(wb):
    headers = {}
    sheet = wb.get_sheet_by_name("Headers")
    for i in range(2,23):
        header = sheet.cell(row=i, column=2).value
        if header not in headers:
            headers[header] = []
        headers[header].append(sheet.cell(row=i, column=1).value)
    return headers

def getPackages(wb):
    sheet = wb.get_sheet_by_name("Packages")
    headers = {}
    for i in range(2,192):
        package = {}
        name = sheet.cell(row=i, column=1).value
        header = sheet.cell(row=i, column=2).value
        package['name'] = name
        package['description'] = sheet.cell(row=i, column=14).value
        site = sheet.cell(row=i, column=13).value
        package['site'] = site if site else "No Source"
        package['keywords'] = {}
        after = sheet.cell(row=i, column=6).value
        if after:
            after = ' ' + after
            if len(after.split(",")) > 1:
                after = ' (' + ' '.join(list(map(lambda x: x.strip(), after.split(",")))) + ')'
            package['keywords']['after'] = after
        package['keywords']['hook'] = sheet.cell(row=i, column=7).value
        package['keywords']['disabled'] = ' t'
        package['keywords']['diminish'] = ''
        if header not in headers:
            headers[header] = []
        headers[header].append(package)
    
    return headers

def createOrgPackageStr(package):
    strList = []
    def app(string):
        strList.append(string)
    app("*** " + package['name'])
    app("#+BEGIN_QUOTE")
    app(package['description'])
    app("#+END_QUOTE")
    app("Source: [[" + package['site'] + "]]")
    app("#+BEGIN_SRC emacs-lisp :tangle ~/.emacs.d/init.el")
    app(createUsePackageStr(package))
    app("#+END_SRC")
    return '\n'.join(strList)

def mergePackage(wb_package, dir_package):
    if dir_package is None:
        return wb_package
    for key in [key for key in wb_package if key != "keywords"]:
        dir_package[key] = wb_package[key]
    
    for key in [key for key in wb_package['keywords']]:
        dir_package['keywords'][key] = wb_package['keywords'][key]
    return dir_package
    
def createOrg(config_root, wb_name, out_file):
    dir_packages = collectAllPackages(config_root)
    wb = load_workbook(wb_name)
    headers = getTopHeaders(wb)
    subheaders = getPackages(wb)
    outString = "#+TITLE: My emacs config\n"
    for header in ["Setup", "Critical Packages", "General Packages", "Language Packages", "Config"]:
        outString += "* " + header + '\n'
        if header in headers:
            for subheader in headers[header]:
                outString += "** " + subheader + '\n'
                for wb_package in subheaders[subheader]:
                    try:
                        dir_package = dir_packages[wb_package['name']]
                    except KeyError:
                        dir_package = None
                    comb_package = mergePackage(wb_package, dir_package)
                    outString += createOrgPackageStr(comb_package) + '\n'

    with open(out_file,"w") as fh:
        fh.write(outString)

        
def main():
    config_root = "./configs"
    wb_name = 'packages.xlsx'
    out_file = "init.org"
    createOrg(config_root, wb_name, out_file) 
    # dir_packages = collectAllPackages(config_root)
    # print(dir_packages['flyspell'])

main()
